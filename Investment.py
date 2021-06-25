import openpyxl as pyxl
from tkinter import *
from tkinter import ttk



root = Tk()
root.title("My investment")
wb = pyxl.load_workbook("try.xlsx")
sh = wb['Investments']
inv = Label (root,text ="Investment:",font = ("Times New Roman" , 14))
inv.grid(row = 0 , column = 0)

invest = ttk.Combobox(root,text = 0)
invest['values'] = (1000,5000,10000,20000,50000,100000)
invest.grid(row=0,column=1)


dte = Label (root,text ="Date:",font = ("Times New Roman" , 14))
dte.grid(row = 1 , column = 0)

date = Spinbox(root,from_=1,to =31,width = 20)
date.grid(row =1 , column = 1)



mnth = Label (root,text ="Month:",font = ("Times New Roman" , 14))
mnth.grid(row = 1 , column = 3)

month = Spinbox(root,from_=1,to =12,width = 20)
month.grid(row =1 , column = 4)


yr = Label (root,text ="Year:",font = ("Times New Roman" , 14))
yr.grid(row = 1 , column = 6)

year = Spinbox(root,from_=2000,to =2025,width = 20)
year.grid(row =1 , column = 7)

inter = Label (root,text ="Interest percentage:",font = ("Times New Roman" , 14))
inter.grid(row = 2 , column = 0)

interest = ttk.Combobox(root)
interest['values'] = (10,20,30,40,50,60,70,80,90,100)
interest.grid(row=2,column=1)

m = 7


def show_calc():

    ip = int(invest.get())*int(interest.get())
    amt = int(ip)/100
    matamt = int(invest.get()) + amt
    matyr = int(year.get()) +1
    mat = Label(root, text = f"Your matured amount = {matamt}/-",font = ("Times New Roman" , 14))
    mat.grid(row=5,column=0)
    matdte = Label(root,text = f"Date: {date.get()}   Month: {month.get()}   Year: {matyr}",font = ("Times New Roman" , 14))
    matdte.grid(row = 6 , column = 0)
    last_row = sh.max_row + 1
    sh.cell(row = last_row ,column = 1, value = str(f"{date.get()}/{month.get()}/{year.get()}"))
    sh.cell(row = last_row ,column = 2, value = invest.get())
    sh.cell(row = last_row,column = 3, value =  interest.get())
    sh.cell(row = last_row,column = 4, value =  str(f"{date.get()}/{month.get()}/{matyr}"))
    sh.cell(row = last_row,column = 5, value =  amt)
    sh.cell(row = last_row,column = 6, value =  matamt)


@show_calc()
def present():

    show_sheet = Tk()
    show_sheet.title("Excel Sheet")
    show_sheet.geometry('500x500')
    sheet_frame = Frame(show_sheet).pack(pady=20)
    sheet_treeview = ttk.Treeview(sheet_frame)
    sheet_treeview['columns'] = ('date', 'amt', 'interests', 'matdte', 'interest_amt', 'matamt')

    sheet_treeview.column('#0', width=0)
    sheet_treeview.column('date', anchor=W, width=120)
    sheet_treeview.column('amt', anchor=W, width=100)
    sheet_treeview.column('interests', anchor=W, width=50)
    sheet_treeview.column('matdte', anchor=E, width=120)
    sheet_treeview.column('interest_amt', anchor=E, width=80)
    sheet_treeview.column('matamt', anchor=E, width=100)

    sheet_treeview.heading('date', text='Date', anchor=W)
    sheet_treeview.heading('amt', text='Deposited Amount', anchor=W)
    sheet_treeview.heading('interests', text='Interest', anchor=W)
    sheet_treeview.heading('matdte', text='Maturity Date', anchor=E)
    sheet_treeview.heading('interest_amt', text='Interest Total Amount', anchor=E)
    sheet_treeview.heading('matamt', text='Maturity Amount', anchor=E)

    sheet_treeview.insert(parent='', index='end', iid=0, values=())





bt_proceed = Button(root,text = "Proceed",font = ('Times New Roman',14), command = show_calc)
bt_proceed.grid(row=3,column = 7)

bt_show = Button(root, text = "Show excel sheet", font = ('Times New ROman',14), command = present)
bt_show.grid(row = 3, column = 5)

root.geometry('700x500')


root.mainloop()
wb.save("try.xlsx")
wb.close()
